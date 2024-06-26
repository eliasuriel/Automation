import pandas as pd
import numpy as np
import os
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import NamedStyle
from openpyxl.worksheet.table import Table, TableStyleInfo
 
import tkinter as tk
from tkinter import filedialog
import easygui


easygui.msgbox("Welcome. Choose your Excel files with PRIME and OPX hours.")
#print("Asking for input file")
root = tk.Tk()
root.withdraw()

easygui.msgbox("Choose your Excel file with Prime data")
file_path = filedialog.askopenfilename()
easygui.msgbox("Choose your Excel file with OPX data")
file_path_opx = filedialog.askopenfilename()
#file_path = "C:\\Users\\VEU1GA\\Documents\\Visual_Studio\\Script2"

excel_file = pd.ExcelFile(file_path)
opx_file = pd.ExcelFile(file_path_opx)

######## HACER CAMBIOS EN ARCHIVO OPX PARA CELDAS COMBINADAS #######

hoja_nombre = opx_file.sheet_names[0]

# Obtener el DataFrame de la hoja especificada
df_2 = opx_file.parse(hoja_nombre)

# Crear un objeto openpyxl Workbook a partir del DataFrame
workbook_de_paso = openpyxl.Workbook()
hoja_prueba = workbook_de_paso.active

# Copiar los datos del DataFrame a la hoja de openpyxl
for i, row in enumerate(df_2.values):
    for j, value in enumerate(row):
        hoja_prueba.cell(row=i+1, column=j+1, value=value)

# Descombinar las celdas en la hoja
for rango in hoja_prueba.merged_cells.ranges:
    # Obtener el valor de la celda combinada
    valor_celda_combinada = hoja_prueba[rango.start_cell].value
    
    # Iterar sobre cada celda en el rango y asignarle el mismo título
    for celda in rango.cells:
        hoja_prueba[celda].value = valor_celda_combinada

celda1 = hoja_prueba['A1']
celda2 = hoja_prueba['B1']
celda = hoja_prueba.cell(row=2, column=1, value=celda1.value)
celda = hoja_prueba.cell(row=2, column=2, value=celda2.value)
hoja_prueba.delete_rows(1)


celda1 = hoja_prueba['C1']
if '202' in celda1.value:
    for i in range(11):
        celda = hoja_prueba.cell(row=1, column=4+i, value=celda1.value +"."+str(i+1))

celda2 = hoja_prueba['O1']
if '202' in celda1.value:
    for i in range(11):
        celda = hoja_prueba.cell(row=1, column=16+i, value=celda2.value +"."+str(i+1))

archivos = str(easygui.enterbox(msg="Write a name to the modified OPX file"))  
archivo  = archivos + ".xlsx"
easygui.msgbox("Enter the folder where you want the modified OPX file to be stored ")
directorio=str(easygui.diropenbox()) 
ruta_completa = os.path.join(directorio, archivo)
workbook_de_paso.save(ruta_completa)

opx_file = pd.ExcelFile(ruta_completa)

##### CONTINUAR CON EL PROCESO ##########3
easygui.msgbox("The files were processed successfully")

anio =  int(easygui.enterbox("Type the year you want to work on"))
str_anio = str(anio)

numhojas = 0


hojas = []
hojas_2 = []


  

#PRIME#
#hojas.append("Associate_Jan")
#hojas.append("Associate_Feb")
#hojas.append("Associate_March")
#hojas.append("Associate_April")
#hojas.append("Associate_May")
#hojas.append("Associate_June")
#hojas.append("Associate_July")

#OPX
hojas_2 = opx_file.sheet_names

#We are going to use only two names, because is the same name of the two columns for all the sheets
#nombrecolumna1 = str(easygui.enterbox("Name of the first column to work on? (coworkers,name of the jobs)"))
#nombrecolumna2 = str(easygui.enterbox("Name of the second column to work on? (hours)"))

#name of the columns of PRIME file
nombrecolumna1 ="Resource Name"
nombrecolumna2= "Hours"

Nombres_Tareas = []
hours = []
column_headers = []
seleccion_hoja = []
Proyectos = []



cont = 0
cont_anio = 0
suma_prime = 0
anio_mes = ""
palabra = "BE-"
palabra2 = "PN-"
condicional = 0
suma_opx = 0
totalopx = 0
mesopx = ""
res_prime = []
res_opx = []
Nombres = ''



###########################################################
#                   Archivo Output
# #########################################################   
 
archivos = str(easygui.enterbox(msg="Write a name to the output file "))  
archivo  = archivos + ".xlsx"
 

easygui.msgbox("Enter the folder where you want the output file to be stored ")
directorio=str(easygui.diropenbox())


 
ruta_completa = os.path.join(directorio, archivo)
print(ruta_completa)
 
# Verificar si el archivo existe
 
if not os.path.exists(ruta_completa):
    workbook = openpyxl.Workbook()      #Crear un nuevo libro
    hoja = workbook.active            #Crear una nueva hoja
    hoja.title = str_anio               #Título de la hoja = anio
    hoja2 = workbook.create_sheet(title="Lista asociados " + str_anio)             
    #hoja2.title = "Lista asociados " + str_anio
    workbook.save(ruta_completa)        #Guardar el libro
 
    easygui.msgbox(f"The file {archivo} was created in: {directorio}")
    
else:
    easygui.msgbox(f"The file {archivo} already exists in: {directorio}")
    workbook = openpyxl.load_workbook(ruta_completa)
    hoja = workbook.create_sheet              #Crear una nueva hoja
    hoja.title = str_anio               #Título de la hoja = anio
    hoja2 = workbook.create_sheet             
    hoja2.title = "Lista asociados " + str_anio
    workbook.save(ruta_completa)        #Guardar el libro



HOURS_PER_DAY =  float(easygui.enterbox("How many laboral hours does a day have? (8, 8.25...)"))
#HOURS_PER_DAY = 8.25

num_fila = 1
num_columna = 1



#########           Estilos de Celdas      ##################
# Definir los colores de relleno
relleno_titulos = PatternFill(start_color="77bedc", end_color="77bedc", fill_type="solid")

#Defenir el estilo de letras
negritas = Font(bold = True)

celda = hoja["B1"] 
celda.value = "Hours/day"
celda.fill = relleno_titulos
celda.font = negritas

celda = hoja["C1"]
celda.value = HOURS_PER_DAY
celda.fill = relleno_titulos
celda.font = negritas





#########################################################
#             Iteraciones y Calculos
#######################################################

area = ["CSW", "NET", "DCOM", "DSW", "TST", "SYS", "ASW"]
proyectos = ["BRP", "Zoox", "Faraday","Tesla", "Ford", "Singer", "Harley Davidson", "GM", "Lordstwon", "Oshkosh", "Lucid", "Navistar", "Rexroth", "METALSA", "ASNA", "ReCar", "Fisker","Battle_Motors","BYD"]

gente_sin_egb = [
    "Humberto Chávez Chávez",
    "Ricardo Adolfo del Razo Real",
    "Atlahuac", 
    "Aguirre Vivo Mayra Fernanda", 
    "Diaz Paredes Jose Ronaldo",
    "Perez Mejia Marcelo Armando", 
    "Dominguez Barba Jose Ramon", 
    "Centeno Alcaraz Jaime Alberto",
    "Perez Fonseca Erick Zahid"
    ]

num_proyectos = int(easygui.enterbox("How many projects do you want to work on?"))

num_columna_2 = 2
num_fila_2 = 2

Nombres = ''

celda = hoja2.cell(row=num_fila_2, column=num_columna_2, value="Project")
celda.fill = relleno_titulos
celda = hoja2.cell(row=num_fila_2, column=num_columna_2+1, value="Month")
celda.fill = relleno_titulos
celda = hoja2.cell(row=num_fila_2, column=num_columna_2+2, value="Component")
celda.fill = relleno_titulos
celda = hoja2.cell(row=num_fila_2, column=num_columna_2+3, value="Name")
celda.fill = relleno_titulos
celda = hoja2.cell(row=num_fila_2, column=num_columna_2+4, value="Activity")
celda.fill = relleno_titulos
celda = hoja2.cell(row=num_fila_2, column=num_columna_2+5, value="Hours Reported")
celda.fill = relleno_titulos
num_fila_2 += 1



for j in range(num_proyectos):
        
        mes = []
        hojas = []

        seleccion_proyecto = easygui.choicebox("Choose the " + str(j + 1) + "° project do you want to work" , choices=proyectos)
        proyecto = seleccion_proyecto

        numhojas = int(easygui.enterbox("For this project: " + str(proyecto) + "; \nHow many MONTHS would you like to analyze?"))
        #numhojas = 1

        nombres_hojas_prime = excel_file.sheet_names
        

        for i in range(numhojas):
            #nombre_hoja = str(easygui.enterbox("In which Prime file sheet (or month) do you want to work? \nSheet #"  + str(i + 1)))
            seleccion = easygui.choicebox("In which Prime file sheet (or month) do you want to work?" + str(i + 1), choices=nombres_hojas_prime)
            hojas.append(seleccion)
            # Mostrar el resultado
            #easygui.msgbox(f"Seleccionaste: {seleccion}")
                

            while True:
                num_mes = int(easygui.enterbox("For " + str(hojas[i]) + ":\nType the month number that corresponds to it.\n(e.g. January = 1, February = 2, ...etc)"))
                if num_mes < 1 or num_mes > 12:
                    easygui.msgbox(msg="The month number must be a NUMBER between 1 and 12.\nPlease write it again.",title="ERROR", ok_button="I'm Sorry :(")
                else:
                    mes.append(num_mes)
                    break

        
        #Imprimir proyecto
        num_columna = 2
        num_fila += 2
        celda = hoja.cell(row=num_fila, column=num_columna, value=proyecto)
        celda.fill=relleno_titulos
        celda.font=negritas

        #Imprimir areas
        for h in range(len(area)):
                for k in range(3):
                    celda = hoja.cell(row=num_fila, column=(1+k+num_columna+3*h), value=area[h])
                    celda.fill=relleno_titulos
                celda = hoja.cell(row=num_fila+1, column=(1+num_columna+3*h), value="OPX")
                celda.fill = relleno_titulos

                celda = hoja.cell(row=num_fila+1, column=(2+num_columna+3*h), value="Prime")
                celda.fill = relleno_titulos

                celda = hoja.cell(row=num_fila+1, column=(3+num_columna+3*h), value="Relation")
                celda.fill = relleno_titulos
        
        num_fila += 2
   

        for i in range(numhojas):

            
            df = excel_file.parse(hojas[i])
            df_2= opx_file.parse(hojas_2[0])
                    
            df = df.dropna(subset=[nombrecolumna1, nombrecolumna2])
            column_headers = df_2.columns.tolist() #   
            df_2[column_headers] = df_2[column_headers].fillna(0)

            if mes[i] == 1:
                anio_mes = str_anio 
            elif mes[i] == 2:
                anio_mes = str_anio + ".1"
            elif mes[i] == 3:
                anio_mes = str_anio + ".2"
            elif mes[i] == 4:
                anio_mes = str_anio + ".3"
            elif mes[i] == 5:
                anio_mes = str_anio + ".4"
            elif mes[i] == 6:
                anio_mes = str_anio + ".5"
            elif mes[i] == 7:
                anio_mes = str_anio + ".6"
            elif mes[i] == 8:
                anio_mes = str_anio + ".7"
            elif mes[i] == 9:
                anio_mes = str_anio + ".8"
            elif mes[i] == 10:
                anio_mes = str_anio + ".9"
            elif mes[i] == 11:
                anio_mes = str_anio + ".10"
            elif mes[i] == 12:
                anio_mes = str_anio + ".11"
            else:
                anio_mes = "0"


                 
                
            for h in range(len(area)):    

                for col_header in column_headers: #recorre los headers del opx-*89
        
                    if str_anio in str(col_header) and anio_mes == str(col_header): #verifica si coincide 2023(o el anio que escriban) o si esta en algun header
                    #cont_anio += 1
                        for index, row in df.iterrows():
                                columna_tareas = row[nombrecolumna1]
                                columna_horas = row[nombrecolumna2]    
                     
                                
                                if ('EGB' in columna_tareas) or (columna_tareas in gente_sin_egb):# or "P_" not in columna_tareas or "N_" not in columna_tareas or "Other" not in columna_tareas:
                                    Nombres = columna_tareas
                                  
                                    
                            
                                if proyecto in str(columna_tareas) and area[h] in str(columna_tareas):
                                    suma_prime = suma_prime + columna_horas
                                    Nombres_Tareas.append(Nombres)
                                    celda = hoja2.cell(row=num_fila_2, column=num_columna_2, value=proyecto)
                                    celda = hoja2.cell(row=num_fila_2, column=num_columna_2+1, value=hojas[i])
                                    celda = hoja2.cell(row=num_fila_2, column=num_columna_2+2, value=area[h])
                                    celda = hoja2.cell(row=num_fila_2, column=num_columna_2+3, value=Nombres)
                                    celda = hoja2.cell(row=num_fila_2, column=num_columna_2+4, value=columna_tareas)
                                    celda = hoja2.cell(row=num_fila_2, column=num_columna_2+5, value=columna_horas)
                                    #Nombres = " "
                                    num_fila_2 += 1
                                    





                        for index, row in df_2.iterrows():                    
                            columna_1 = row[column_headers[0]]
                            columna_2 = row[col_header] 
                            if index == 0:
                                #easygui.msgbox(columna_2)  
                                mesopx = columna_2
                                celda = hoja.cell(row=num_fila, column=num_columna, value=mesopx)
                                celda.fill = relleno_titulos                     
                            
                            #print(columna_1)
                            #print(columna_2)

                            if str(columna_1).startswith(palabra or palabra2):
                                condicional = 0
                            
                            if proyecto in str(columna_1):
                                condicional = 1
                            
                            if condicional == 1 and area[h] in str(columna_1):
                                suma_opx = suma_opx + columna_2*HOURS_PER_DAY     

                #easygui.msgbox("Mes Prime: " + str(hojas[i]) + "\nMes OPX: " + str(mesopx) + "\nProyecto: " + str(proyecto) + "\nArea: " + str(area[h])  + "\nHoras Prime: " + str(suma_prime) + "\nHojas OPX: " + str(suma_opx))
                
                resultados = []
                res_prime.append(suma_prime)
                res_opx.append(suma_opx) 
                resultados.append(suma_opx)
                resultados.append(suma_prime)
                if suma_opx == 0:
                    relacion = 0
                else:
                    relacion = suma_prime/suma_opx
                resultados.append(relacion)

                for k in range(3):
                    celda = hoja.cell(row=num_fila, column=(1+k+num_columna+3*h), value=resultados[k])

                if suma_prime !=0:
                    suma_prime = 0
                if suma_opx !=0:
                    suma_opx = 0
            num_fila += 1



######### HACER TABLA HOJA 2 #################
existing_tables = hoja2.tables
table_name = 'Tabla_general'
 
if table_name in existing_tables:
    # Si la tabla ya existe, la eliminamos
    hoja2._tables.remove(existing_tables[table_name])
 
 
#Crea una nueva tabla
table = Table(displayName=table_name, ref=("B2:G")+str(num_fila_2 - 1))
style = TableStyleInfo(
    name="TableStyleMedium9", showFirstColumn=False,
    showLastColumn=False, showRowStripes=False, showColumnStripes=True)
table.tableStyleInfo = style
hoja2.add_table(table)




############# Ajustar ancho de columnas ####################
# Ajustar el ancho de las columnas de la tabla
for columna in hoja.iter_cols(min_col=2, max_col=23):
    longitud_maxima = 0
    columna_letra = columna[0].column_letter  # Obtiene la letra de la columna
    for celda in columna:
        if celda.value:
            if celda.data_type == 'n':
                longitud_celda = 8
            else:
                # Calcula la longitud máxima del contenido en la columna
                longitud_celda = len(str(celda.value))
            
            if longitud_celda > longitud_maxima:
                longitud_maxima = longitud_celda
 
    # Establece el ancho de la columna para ajustarlo al contenido más largo
    hoja.column_dimensions[columna_letra].width = longitud_maxima +2

# Ajustar el ancho de las columnas de la tabla 2
for columna in hoja2.iter_cols(min_col=2, max_col=7):
    longitud_maxima = 0
    columna_letra = columna[0].column_letter  # Obtiene la letra de la columna
    for celda in columna:
        if celda.value:
            if celda.data_type == 'n':
                longitud_celda = 8
            else:
                # Calcula la longitud máxima del contenido en la columna
                longitud_celda = len(str(celda.value))
            
            if longitud_celda > longitud_maxima:
                longitud_maxima = longitud_celda
 
    # Establece el ancho de la columna para ajustarlo al contenido más largo
    hoja2.column_dimensions[columna_letra].width = longitud_maxima +4


#Guarda el archivo
workbook.save(ruta_completa)
 
#Abre el archivo
if os.path.exists(ruta_completa):
    easygui.msgbox("Save completed, opening file...")
    os.startfile(ruta_completa)
    #os.system(f'start excel "{ruta_completa}"')
else:
    easygui.msgbox(f'El archivo "{ruta_completa}" no existe.')
